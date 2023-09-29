from flask import Flask, render_template, url_for, request, send_file
import pandas as pd
from io import BytesIO 
from reportlab.pdfgen import canvas
import openpyxl
import csv
from openpyxl import Workbook,load_workbook



app = Flask(__name__)

#Adding File Reading Function:
def xread(path):

    path=path.strip()

    if (path[-3:]=="csv"):
        with open(path, 'r') as f:
            reader = csv.reader(f)
            metadata = list(reader)
        return metadata

    elif (path[-4:]=="xlsx"):
        xl=load_workbook(path)
        sheet=xl.active
        maxrow=sheet.max_row
        maxcol=sheet.max_column
        metadata=[]
        mtdt=[]

        for j in range(1, maxrow + 1): #to replace into column wise data set.. change to maxrow+1 to maxcol+1
            for i in range(1, maxcol + 1): #vice versa
                cell = sheet.cell(row=j, column=i) #switch i and j
                mtdt.append(cell.value)
            metadata.append(mtdt)
            mtdt=[]
        return metadata

    else:
        return "Error. File not found or Invalid extension type. Please recheck path and use only csv or xlsx files."



@app.route('/', methods = ['GET', 'POST'])
def index():
    if request.method == "POST":
        csv_file = request.files['csv_file']
        if csv_file:
            df = pd.read_csv(csv_file)
            pdf = generate_pdf(df)
            return send_file(pdf, as_attachment=True, download_name='output.pdf', mimetype='application/pdf')
    return render_template("index.html")

def generate_pdf(df):
    buffer = BytesIO()
    p = canvas.Canvas(buffer)
    p.save()
    buffer.seek(0)
    return buffer

if __name__ == "__main__":
    app.run(debug=True)
